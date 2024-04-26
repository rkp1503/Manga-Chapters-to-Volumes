"""
Author: Rayla Kurosaki

GitHub: https://github.com/rkp1503
"""

import copy
import math
import os
import shutil
import sys

from openpyxl import load_workbook

import web_scraper as ws
from comicinfo import ComicInfo

# List of manga that is presorted by volume
PRESORTED: list = [
    "King's Game", "Mayo Chiki!", "Strobe Edge",
    "",
    "Assorted Entanglements", "Citrus+",
    "I Think I Turned My Childhood Friend Into a Girl", "Pokémon Adventures"
]


def get_volume_chapter_data(path_to_manga: str) -> dict:
    # Store the chapter numbers in a volume via a temp dictionary
    vol_ch_data: dict = {}
    # Get the path to the data workbook
    path_to_data: str = os.path.join(path_to_manga, "data.xlsx")
    # Parse through the "Book" sheet in the data workbook
    for (i, row) in enumerate(load_workbook(path_to_data)["Book"].iter_rows()):
        if i > 0:
            # Store the data in a dictionary
            vol_num: str = str(row[1].value)
            first_ch: float = float(row[4].value)
            last_ch: float = float(row[5].value)
            vol_ch_data[vol_num] = (first_ch, last_ch)
            pass
        pass
    return vol_ch_data


def create_volume_directories(path_to_manga: str, vol_ch_data: dict) -> None:
    # For each volume
    for vol_num in vol_ch_data.keys():
        # Get the path to the folder
        path_to_vol: str = os.path.join(path_to_manga, f"Volume {vol_num}")
        # If the ".cbz" file (or folder) does not exist
        if not os.path.exists(path_to_vol):
            if not os.path.exists(f"{path_to_vol}.cbz"):
                # Create the folder
                os.mkdir(path_to_vol)
                pass
            pass
        pass
    return None


def get_padding(curr: float, last: float) -> str:
    if curr < 1:
        curr = 1
        pass
    return "0" * (math.floor(math.log10(last)) - math.floor(math.log10(curr)))


def fill_volume_directories(path_to_manga: str, vol_ch_data: dict) -> None:
    # For each chapter in the manga
    for chapter in os.listdir(path_to_manga):
        if (".xlsx" not in chapter) and ("Volume" not in chapter):
            # Get the chapter number as a number
            ch_num: float = float(chapter.split(" ")[-1])
            # Convert the number to a string and add padding
            ch_pad: str = f"{get_padding(ch_num, 1000)}{ch_num}"
            # Remove the "."
            pg_pref: str = f"{ch_pad.split('.')[0]}{ch_pad.split('.')[1]}"
            # Get the path to the chapter
            path_to_ch: str = os.path.join(path_to_manga, chapter)
            # Find the volume the chapter belongs to
            vol_num: str = ""
            for (k, v) in vol_ch_data.items():
                if v[0] <= ch_num <= v[1]:
                    vol_num = k
                    break
                    pass
                pass
            # Get the path to the correct volume directory
            path_to_vol: str = os.path.join(path_to_manga, f"Volume {vol_num}")
            # For each page in the chapter
            for pg in os.listdir(path_to_ch):
                if (pg != ".nomedia") and (pg != "ComicInfo.xml"):
                    # Get the source path for the page
                    src: str = os.path.join(path_to_ch, pg)
                    # Get the destination path for the page
                    dst: str = os.path.join(path_to_vol, f"{pg_pref}_{pg}")
                    # Move and rename the page in the volume directory
                    os.rename(src, dst)
                    pass
                pass
            # Remove the chapter directory
            shutil.rmtree(path_to_ch)
            pass
        pass
    return None


def rename_pages(path_to_manga: str) -> None:
    # For each volume directory
    for volume in os.listdir(path_to_manga):
        bool_1: bool = ".xlsx" not in volume
        bool_2: bool = ".cbz" not in volume
        bool_3: bool = volume != "Volume Covers"
        if bool_1 and bool_2 and bool_3:
            # Get the path to the volume directory
            path_to_vol: str = os.path.join(path_to_manga, volume)
            # Get the number of pages in the volume directory
            pg_count: int = len(os.listdir(path_to_vol))
            # For each page in the volume
            for (i, page) in enumerate(os.listdir(path_to_vol)):
                # Get the original file extension
                ext: str = page.split(".")[-1]
                # Get the path to the page
                src: str = os.path.join(path_to_vol, page)
                # Get a new name for the page
                new_name: str = f"{get_padding(i + 1, pg_count)}{i + 1}.{ext}"
                # Get the path to the renamed page
                dst: str = os.path.join(path_to_vol, new_name)
                # Rename the page
                os.rename(src, dst)
                pass
            pass
        pass
    return None


def create_volume_cover_directory(path_to_manga: str) -> None:
    vol_covers_path: str = os.path.join(path_to_manga, "Volume Covers")
    if not os.path.exists(vol_covers_path):
        os.mkdir(vol_covers_path)
        pass
    return None


def fill_volume_cover_directory(path_to_manga: str) -> None:
    # For each volume directory
    for volume in os.listdir(path_to_manga):
        bool_1: bool = ".xlsx" not in volume
        bool_2: bool = ".cbz" not in volume
        bool_3: bool = volume != "Volume Covers"
        if bool_1 and bool_2 and bool_3:
            # Create the path to the volume
            path_to_vol: str = os.path.join(path_to_manga, volume)
            # Get the first page of the volume
            pg: str = os.listdir(path_to_vol)[0]
            # Get the path to the first page of the volume
            src: str = os.path.join(path_to_vol, pg)
            # Get the path to the cover directory
            path_to_covers: str = os.path.join(path_to_manga, "Volume Covers")
            # Get the extension of the copied/original file
            ext: str = pg.split(".")[-1]
            # Get the volume number
            vol_num: str = volume.split(" ")[-1]
            # Get the new name for the volume cover
            new_name: str = f"Volume Cover {vol_num}.{ext}"
            # Get the path to the destination location
            dst: str = os.path.join(path_to_covers, new_name)
            # Copy the file to the new location
            shutil.copy(src, dst)
            pass
        pass
    return None


def add_metadata(path_to_manga: str) -> None:
    print(f"\tGetting metadata...")
    url_BUM: str = ""
    # Create a metadata object
    metadata: ComicInfo = ComicInfo()
    # Get the path to the data workbook
    path_to_data: str = os.path.join(path_to_manga, "data.xlsx")
    # Parse through the "Series" sheet in the data workbook
    for row in load_workbook(path_to_data)["Series"].iter_rows():
        category: str | None = row[0].value
        data: str | None = row[1].value
        if (category is not None) and (data is not None):
            if category == "Source":
                pass
            # Add manga details from Baka Updates Manga
            elif category == "Baka-Updates Manga":
                url_BUM = data
                ws.baka_updates_manga(data, metadata)
                pass
            # Add manga details from My Anime List
            elif category == "MyAnimeList":
                ws.my_anime_list(data, metadata)
                pass
            # Add manga details from Wikipedia
            elif category == "Wikipedia":
                ws.web_scraper_wiki(data, metadata)
                pass
            # Set the manga's status
            elif category == "Status":
                metadata.set_status(data)
                # Set the number of books if the series is not an ongoing
                # series
                if data in ["Completed", "Ended", "Abandoned", "Cancelled"]:
                    count: int = len(os.listdir(path_to_manga)) - 2
                    metadata.set_count(count)
                    pass
                pass
            elif category == "Collection":
                metadata.add_groups(data.split(", "))
                pass
            elif category == "Read List":
                metadata.set_story_arc(data)
                pass
            elif category == "Notes":
                metadata.set_notes(data)
                pass
            pass
        pass
    # Parse through the "Book" sheet in the data workbook
    for (i, row) in enumerate(load_workbook(path_to_data)["Book"].iter_rows()):
        if i > 0:
            title: str = row[0].value
            # Get the path to the corresponding volume directory
            path_to_vol: str = os.path.join(path_to_manga, title)
            if os.path.exists(path_to_vol):
                print(f"\t\t{title}...")
                # Create a hard copy of the metadata with the series details
                metadata_temp: ComicInfo = copy.deepcopy(metadata)
                # Set the book title
                metadata_temp.set_title(title)
                # Set the sorting number for the series
                number: str = row[2].value
                metadata_temp.set_number(number)
                # Set the sorting number for the read list if provided
                read_list_number: str | None = row[3].value
                if read_list_number is not None:
                    metadata_temp.set_story_arc_number(read_list_number)
                    pass
                # Set the url to the book's source
                url: str = row[6].value
                metadata_temp.add_url(url)
                # Extract and store the book's metadata from the give url
                ws.web_scraper_publisher(url, metadata_temp)
                # Add sports tag
                metadata_temp.add_sports_tag()
                # If the series title is not set in the original metadata
                # object
                if metadata.get_series_title() == "":
                    # Set the series title in the original metadata object
                    metadata.set_series_title(metadata_temp.get_series_title())
                    # Check for any conflicting genres in the hard copied
                    # metadata object
                    metadata_temp.check_conflicting_genres_and_tags()
                    # Set the genres in the original metadata object
                    metadata.set_genres(metadata_temp.get_genres())
                    # Set the tags in the original metadata object
                    metadata.set_tags(metadata_temp.get_tags())
                    # print(f"\t\tGenres: {metadata.get_genres()}")
                    # print(f"\t\tTags: {metadata.get_tags()}")
                    pass
                # Create the path to a writable text file
                path_to_txt: str = os.path.join(path_to_vol, "ComicInfo.txt")
                with open(path_to_txt, "w", encoding="utf-8") as file:
                    file.write(metadata_temp.get_xml_data())
                    pass
                # Get the path to the metadata file
                path_to_xml: str = os.path.join(path_to_vol, "ComicInfo.xml")
                # Rename the file extension from ".txt" to ".xml"
                os.rename(path_to_txt, path_to_xml)
                pass
            pass
        pass
    # print(f"\t\t{metadata.get_genres()}")
    return None


def convert_volume_directories_to_cbz(path_to_manga: str) -> None:
    print(f"\tConverting to cbz...")
    # For each volume directory
    for volume in os.listdir(path_to_manga):
        bool_1: bool = ".xlsx" not in volume
        bool_2: bool = ".cbz" not in volume
        bool_3: bool = volume != "Volume Covers"
        if bool_1 and bool_2 and bool_3:
            print(f"\t\t{volume}")
            # Get the path to the volume directory
            path_to_vol: str = os.path.join(path_to_manga, volume)
            # Zip the volume directory
            shutil.make_archive(volume, "zip", path_to_vol)
            # Get the path to the zip file
            path_to_zip: str = f"{os.path.join(os.getcwd(), volume)}.zip"
            # Get the path to the ".cbz" file
            path_to_cbz: str = f"{path_to_vol}.cbz"
            # Change the extension from .zip to .cbz
            os.rename(path_to_zip, path_to_cbz)
            # Remove the volume directory
            shutil.rmtree(path_to_vol)
            pass
        pass
    return None


def get_vol_count(vol_ch_data: dict) -> float:
    vol_count: float = -1
    for i in vol_ch_data.keys():
        for j in i.split(" ")[-1].split("-"):
            if float(j) > vol_count:
                vol_count = float(j)
                pass
            pass
        pass
    return vol_count


def rename_volume_directories(path_to_manga: str, vol_ch_data: dict) -> None:
    max_vol: float = get_vol_count(vol_ch_data)
    # Only rename directories if the volume count is at least 10
    if max_vol >= 10:
        # For each volume directory
        for volume in os.listdir(path_to_manga):
            bool_1: bool = ".xlsx" not in volume
            bool_2: bool = ".cbz" not in volume
            bool_3: bool = volume != "Volume Covers"
            if bool_1 and bool_2 and bool_3:
                # Get the path to the current directory
                vol_src: str = os.path.join(path_to_manga, volume)
                # Get the volume number
                vol_num: float = float(volume.split(" ")[-1])
                # Get the padding for the volume number
                vol_num_pad: str = get_padding(vol_num, max_vol)
                # Create the new name for the volume directory
                new_name: str = f"Volume {vol_num_pad}{vol_num}"
                # Get the path to the renamed volume directory
                vol_dst: str = os.path.join(path_to_manga, new_name)
                # Rename the volume directory
                os.rename(vol_src, vol_dst)
                pass
            pass
        pass
    return None


def remove_non_pages(path_to_manga: str) -> None:
    # For each volume directory
    for volume in os.listdir(path_to_manga):
        bool_1: bool = ".xlsx" not in volume
        bool_2: bool = ".cbz" not in volume
        bool_3: bool = volume != "Volume Covers"
        if bool_1 and bool_2 and bool_3:
            # Get the path to the volume directory
            path_to_vol: str = os.path.join(path_to_manga, volume)
            # Remove the ".nomedia" file if it exists
            path_to_file: str = os.path.join(path_to_vol, ".nomedia")
            if os.path.exists(path_to_file):
                os.remove(path_to_file)
                pass
            # Remove the "Comicinfo.xml" file if it exists
            path_to_file: str = os.path.join(path_to_vol, "Comicinfo.xml")
            if os.path.exists(path_to_file):
                os.remove(path_to_file)
                pass
            pass
        pass
    return None


# def add() -> None:
#     return None
#
#
# def update_library() -> None:
#     return None
#
#
# def update_metadata() -> None:
#     return None


def main() -> None:
    # Get the path to the library
    path_to_lib: str = sys.argv[1]
    # Get the command to execute.
    # Valid commands are "-add", "-u lib", "-u data"
    # For each manga in the library
    for manga in os.listdir(path_to_lib):
        print(f"Series \"{manga}\"...")
        # Get the path to the manga
        path_to_manga: str = os.path.join(path_to_lib, manga)
        # If the data workbook exists
        path_to_data: str = os.path.join(path_to_manga, "data.xlsx")
        if os.path.exists(path_to_data):
            # Get the volume to chapter data
            vol_ch_data: dict = get_volume_chapter_data(path_to_manga)
            if manga in PRESORTED:
                # Rename all the volume directories
                rename_volume_directories(path_to_manga, vol_ch_data)
                # Remove all irrelevant items
                remove_non_pages(path_to_manga)
                pass
            else:
                # Create the volume directories
                create_volume_directories(path_to_manga, vol_ch_data)
                # Fill the volume directories
                fill_volume_directories(path_to_manga, vol_ch_data)
                pass
            # Rename the pages in the volume directories
            rename_pages(path_to_manga)
            # Create the volume cover directory
            create_volume_cover_directory(path_to_manga)
            # Fill the volume cover directory
            fill_volume_cover_directory(path_to_manga)
            # Add metadata to each volume
            add_metadata(path_to_manga)
            # Zip the volume directories and convert it to a ".cbz" file
            convert_volume_directories_to_cbz(path_to_manga)
            print(f"Series \"{manga}\" completed!\n")
            pass
        else:
            # Skip the series if the data workbook does not exist
            print(f"Missing \"data.xlsx\" file. Skipping {manga}...")
        pass
    return None


def main_test():
    metadata: ComicInfo = ComicInfo()

    url_bum = "https://www.mangaupdates.com/series/kim0rqm/absolute-duo"
    url_mal = "https://myanimelist.net/manga/51023/Absolute_Duo"
    url_wiki = "https://en.wikipedia.org/wiki/Absolute_Duo"
    url_pubs = [
        # "https://sevenseasentertainment.com/books/absolute-duo-vol-1/",
        "https://sevenseasentertainment.com/books/akuma-no-riddle-riddle-story-of-devil-vol-1/",
        # "https://sevenseasentertainment.com/books/bloom-into-you-vol-1/",
        # "https://sevenseasentertainment.com/books/citrus-vol-1/",
        # "https://sevenseasentertainment.com/books/days-of-love-at-seagull-villa-vol-1/",
        # "https://sevenseasentertainment.com/books/kashimashi-girl-meets-girl-vol-1/",
        # "https://sevenseasentertainment.com/books/love-me-for-who-i-am-vol-1/",
    ]

    # ws.baka_updates_manga(url_bum, metadata)
    # ws.my_anime_list(url_mal, metadata)
    # ws.web_scraper_wiki(url_wiki, metadata)
    for url_pub in url_pubs:
        metadata_temp: ComicInfo = copy.deepcopy(metadata)
        ws.web_scraper_publisher(url_pub, metadata_temp)
        # print(metadata_temp.get_xml_data())
        # print()
        pass
    pass


if __name__ == '__main__':
    # main()
    main_test()
    pass
